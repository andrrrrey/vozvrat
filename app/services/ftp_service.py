"""FTP integration service for 1C data exchange."""
import logging
import ftplib
import asyncio
import io
import os
import csv
from typing import Optional
from datetime import datetime, timezone
from sqlalchemy.ext.asyncio import AsyncSession
import openpyxl

from app.services.settings_service import get_setting, set_setting

logger = logging.getLogger(__name__)


async def _get_ftp_settings(db: AsyncSession) -> dict:
    return {
        "host": await get_setting(db, "ftp_host"),
        "port": int(await get_setting(db, "ftp_port") or "21"),
        "user": await get_setting(db, "ftp_user"),
        "password": await get_setting(db, "ftp_password"),
        "path": await get_setting(db, "ftp_path") or "/",
    }


def _test_ftp_sync(cfg: dict) -> str:
    """Test FTP connection synchronously. Returns success message or raises."""
    ftp = ftplib.FTP()
    ftp.connect(cfg["host"], cfg["port"], timeout=10)
    ftp.login(cfg["user"], cfg["password"])
    welcome = ftp.getwelcome()
    ftp.quit()
    return welcome or "Подключение успешно"


async def test_ftp_connection(db: AsyncSession) -> str:
    cfg = await _get_ftp_settings(db)
    if not cfg["host"]:
        raise ValueError("FTP хост не настроен")
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _test_ftp_sync, cfg)


def _upload_file_sync(cfg: dict, remote_path: str, data: bytes) -> None:
    ftp = ftplib.FTP()
    ftp.connect(cfg["host"], cfg["port"], timeout=15)
    ftp.login(cfg["user"], cfg["password"])
    try:
        # Ensure base directory exists
        base = cfg["path"].rstrip("/")
        if base:
            try:
                ftp.cwd(base)
            except ftplib.error_perm:
                ftp.mkd(base)
                ftp.cwd(base)
        ftp.storbinary(f"STOR {remote_path}", io.BytesIO(data))
    finally:
        ftp.quit()


def _upload_file_to_subdir_sync(cfg: dict, subdirs: list[str], remote_name: str, data: bytes) -> str:
    """Upload a file into nested subdirectories (created if missing) under the base path.

    Returns the remote path used. Каждый каталог из subdirs создаётся при отсутствии.
    """
    ftp = ftplib.FTP()
    ftp.connect(cfg["host"], cfg["port"], timeout=20)
    ftp.login(cfg["user"], cfg["password"])
    try:
        base = cfg["path"].rstrip("/")
        if base:
            try:
                ftp.cwd(base)
            except ftplib.error_perm:
                ftp.mkd(base)
                ftp.cwd(base)
        for part in subdirs:
            try:
                ftp.cwd(part)
            except ftplib.error_perm:
                ftp.mkd(part)
                ftp.cwd(part)
        ftp.storbinary(f"STOR {remote_name}", io.BytesIO(data))
        return f"{'/'.join(subdirs)}/{remote_name}"
    finally:
        ftp.quit()


async def upload_1c_xlsx(db: AsyncSession, refunds: list) -> str:
    """Build the 1C XLSX export and upload it via FTP into Vozvrat1C/OUT.

    Папка Vozvrat1C/OUT создаётся автоматически, если её нет. Возвращает удалённый путь.
    """
    cfg = await _get_ftp_settings(db)
    if not cfg["host"]:
        raise ValueError("FTP не настроен")
    data = build_1c_xlsx(refunds)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    remote_name = f"goodsReturnReport_{timestamp}.xlsx"
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(
        None, _upload_file_to_subdir_sync, cfg, ["Vozvrat1C", "OUT"], remote_name, data
    )


def price_cache_path() -> str:
    """Локальный путь, куда сохраняется последний автоскачанный по FTP прайс."""
    from app.config import settings as cfg
    return os.path.join(cfg.UPLOAD_DIR, "price_ftp_cache.xlsx")


def _download_latest_price_sync(cfg: dict, price_path: str) -> tuple[str, bytes]:
    """Скачать самый свежий файл прайса (XLS/XLSX) из указанной папки FTP.

    Возвращает (имя_файла, содержимое). Если файлов несколько — выбирается
    последний по времени модификации (MDTM), с фолбэком на имя."""
    ftp = ftplib.FTP()
    ftp.connect(cfg["host"], cfg["port"], timeout=30)
    ftp.login(cfg["user"], cfg["password"])
    try:
        base = cfg["path"].rstrip("/")
        if base:
            try:
                ftp.cwd(base)
            except ftplib.error_perm:
                ftp.mkd(base)
                ftp.cwd(base)
        # Перейти во вложенную папку с прайсом, если указана.
        for part in [p for p in (price_path or "").strip("/").split("/") if p]:
            ftp.cwd(part)

        names = ftp.nlst()
        candidates = [os.path.basename(n) for n in names
                      if n.lower().endswith((".xls", ".xlsx"))]
        if not candidates:
            raise ValueError("В указанной папке FTP не найден файл прайса (XLS/XLSX).")

        def _mtime(name: str) -> str:
            try:
                resp = ftp.sendcmd("MDTM " + name)  # "213 20240101120000"
                parts = resp.split()
                return parts[1] if len(parts) > 1 else ""
            except Exception:
                return ""

        candidates.sort(key=lambda n: (_mtime(n), n))
        chosen = candidates[-1]

        bio = io.BytesIO()
        ftp.retrbinary("RETR " + chosen, bio.write)
        return chosen, bio.getvalue()
    finally:
        try:
            ftp.quit()
        except Exception:
            pass


async def download_and_cache_price(db: AsyncSession) -> dict:
    """Скачать прайс по FTP из папки `ftp_price_path` и сохранить локально.

    Записывает метаданные в настройки (`price_ftp_last_name/at/size`).
    Возвращает {filename, size}. Бросает ValueError при отсутствии настроек/файла."""
    cfg = await _get_ftp_settings(db)
    if not cfg["host"]:
        raise ValueError("FTP не настроен")
    price_path = (await get_setting(db, "ftp_price_path")).strip()

    loop = asyncio.get_event_loop()
    filename, data = await loop.run_in_executor(None, _download_latest_price_sync, cfg, price_path)

    path = price_cache_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "wb") as f:
        f.write(data)

    await set_setting(db, "price_ftp_last_name", filename)
    await set_setting(db, "price_ftp_last_at", datetime.now(timezone.utc).isoformat())
    await set_setting(db, "price_ftp_last_size", str(len(data)))
    logger.info(f"Price downloaded from FTP: {filename} ({len(data)} bytes)")
    return {"filename": filename, "size": len(data)}


def _build_1c_csv(refunds: list) -> bytes:
    """Build CSV export for 1C from list of Refund objects."""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([
        "Номер возврата",
        "Номер УПД",
        "ID позиции",
        "ID Заказа",
        "Артикул",
        "Производитель",
    ])
    for refund in refunds:
        upd = refund.upd_number or "" if refund.is_manual else ""
        if refund.items:
            for item in refund.items:
                writer.writerow([
                    refund.display_id,
                    upd,
                    item.position_id or "",
                    refund.order_id or "",
                    item.article,
                    item.brand or "",
                ])
        else:
            writer.writerow([
                refund.display_id,
                upd,
                "",
                refund.order_id or "",
                "",
                "",
            ])
    return output.getvalue().encode("utf-8-sig")


async def upload_1c_export(db: AsyncSession, refunds: list) -> str:
    """Upload 1C export CSV to FTP. Returns remote path."""
    cfg = await _get_ftp_settings(db)
    if not cfg["host"]:
        raise ValueError("FTP не настроен")
    data = _build_1c_csv(refunds)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    remote_name = f"export_1c_{timestamp}.csv"
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, _upload_file_sync, cfg, remote_name, data)
    return remote_name


def build_1c_csv(refunds: list) -> bytes:
    """Public sync wrapper for building CSV without FTP upload."""
    return _build_1c_csv(refunds)


def build_1c_xlsx(refunds: list) -> bytes:
    """Build XLSX export for 1C. The position ID, order ID and article columns are
    stored as text to prevent Excel from converting long numeric strings to
    scientific notation. The "Номер УПД" column is always present but filled only
    for manual returns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Возврат 1С"

    headers = [
        "Номер возврата",
        "Номер УПД",
        "ID позиции",
        "ID Заказа",
        "Артикул",
        "Производитель",
    ]
    ws.append(headers)

    # Columns that must stay as text: C=ID позиции (3), D=ID Заказа (4), E=Артикул (5)
    TEXT_COLS = {3, 4, 5}

    def _append_row(row_values: list) -> None:
        ws.append(row_values)
        r = ws.max_row
        for col_idx in TEXT_COLS:
            cell = ws.cell(row=r, column=col_idx)
            cell.value = str(cell.value) if cell.value is not None else ""
            cell.number_format = "@"

    for refund in refunds:
        upd = refund.upd_number or "" if refund.is_manual else ""
        if refund.items:
            for item in refund.items:
                _append_row([
                    refund.display_id,
                    upd,
                    item.position_id or "",
                    refund.order_id or "",
                    item.article,
                    item.brand or "",
                ])
        else:
            _append_row([
                refund.display_id,
                upd,
                "",
                refund.order_id or "",
                "",
                "",
            ])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
