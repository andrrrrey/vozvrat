import asyncio
import logging
import os
from datetime import datetime
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File, Form
from fastapi.responses import Response, JSONResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.services.auth import get_current_user

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/tables", tags=["tables"])


def _patch_openpyxl_noneset():
    """Some xlsx files contain invalid 'pane' attribute values that openpyxl rejects.
    The error comes from NoneSet.__set__ (a descriptor), not __init__.
    Patch NoneSet.__set__ once to silently coerce unknown values to None.
    (Same workaround as app/routers/refunds.py.)"""
    import openpyxl.descriptors.base as _openpyxl_desc

    if not getattr(_openpyxl_desc, "_noneset_patched", False):
        _orig_noneset_set = _openpyxl_desc.NoneSet.__set__

        def _safe_noneset_set(self, instance, value):
            if value is not None and value != "none" and value not in self.values:
                value = None
            _orig_noneset_set(self, instance, value)

        _openpyxl_desc.NoneSet.__set__ = _safe_noneset_set
        _openpyxl_desc._noneset_patched = True


def _norm(value) -> str:
    """Normalize an article/header value for matching: trim + upper-case."""
    if value is None:
        return ""
    return str(value).strip().upper()


# Заголовки колонок с кодами по умолчанию (если в файле-справочнике их не удалось прочитать).
TNVED_HEADER = "код ТН ВЭД"
OKPD_HEADER = "код ОКПД 2"


def _hkey(value) -> str:
    """Normalize a header for keyword matching: upper-case, drop all whitespace
    (including non-breaking) and dots so 'код ТН ВЭД' == 'КОДТНВЭД'."""
    return "".join(str(value or "").upper().split()).replace("\xa0", "").replace(".", "")


def _classify_codes_header(row):
    """Given a header row (tuple of cell values), locate the article / ТН ВЭД / ОКПД
    columns by keyword. Returns dict with indices and original header texts, or None
    if all three are not present.

    Tolerant of spelling variants: 'код ТН ВЭД' / 'код ТЭНВД', 'код ОКПД 2' / 'код ОКПД2'."""
    art_col = tnved_col = okpd_col = None
    tnved_header = okpd_header = None
    for idx, cell in enumerate(row):
        key = _hkey(cell)
        if not key:
            continue
        if art_col is None and "АРТИКУЛ" in key:
            art_col = idx
        elif okpd_col is None and "ОКПД" in key:
            okpd_col = idx
            okpd_header = str(cell).strip()
        elif tnved_col is None and ("ВЭД" in key or "ТЭНВД" in key or "ТНВД" in key or "ВЕД" in key):
            tnved_col = idx
            tnved_header = str(cell).strip()
    if art_col is None or tnved_col is None or okpd_col is None:
        return None
    return {
        "art_col": art_col,
        "tnved_col": tnved_col,
        "okpd_col": okpd_col,
        "tnved_header": tnved_header or TNVED_HEADER,
        "okpd_header": okpd_header or OKPD_HEADER,
    }


def _find_price_header_row(ws, target_norm, max_scan=25):
    """Find the price header row containing target_norm (e.g. 'КАТАЛОЖНЫЙ НОМЕР').
    Returns (row_index, column_index) or (None, None)."""
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        for col_idx, cell in enumerate(row):
            if _norm(cell) == target_norm:
                return row_idx, col_idx
    return None, None


def _build_codes_map(codes_bytes: bytes):
    """Read the codes reference file into {norm(article): (tnved, okpd2)}.
    Returns (codes_map, tnved_header, okpd_header, layout), where layout is
    'header' (columns found by name) or 'positional' (no header row — standard
    column order A/C/D assumed)."""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(BytesIO(codes_bytes), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="Не удалось открыть файл «таблица с кодами». Убедитесь, что это корректный XLS/XLSX.",
        )

    ws = wb.active

    header_row = None
    cols = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=25, values_only=True), start=1):
        cols = _classify_codes_header(row)
        if cols is not None:
            header_row = row_idx
            break

    layout = "header"
    if cols is None:
        # Файл без строки заголовков — используем стандартный порядок колонок:
        # A=артикул, B=бренд, C=код ТН ВЭД, D=код ОКПД 2. Данные с первой строки.
        if ws.max_column is None or ws.max_column < 4:
            wb.close()
            raise HTTPException(
                status_code=400,
                detail=(
                    "В файле «таблица с кодами» не найдены колонки «артикул», «код ТН ВЭД», "
                    "«код ОКПД 2», и файл не подходит под стандартный порядок колонок "
                    "(A — артикул, B — бренд, C — код ТН ВЭД, D — код ОКПД 2)."
                ),
            )
        cols = {
            "art_col": 0,
            "tnved_col": 2,
            "okpd_col": 3,
            "tnved_header": TNVED_HEADER,
            "okpd_header": OKPD_HEADER,
        }
        header_row = 0  # данные начинаются со строки 1
        layout = "positional"

    art_col = cols["art_col"]
    tnved_col = cols["tnved_col"]
    okpd_col = cols["okpd_col"]

    codes_map: dict = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if art_col >= len(row):
            continue
        key = _norm(row[art_col])
        if not key:
            continue
        tnved = row[tnved_col] if tnved_col < len(row) else None
        okpd = row[okpd_col] if okpd_col < len(row) else None
        codes_map[key] = (tnved, okpd)

    wb.close()
    return codes_map, cols["tnved_header"], cols["okpd_header"], layout


def _enrich_price(price_bytes: bytes, codes_map: dict,
                  tnved_header: str = TNVED_HEADER, okpd_header: str = OKPD_HEADER):
    """Load the price workbook, append two code columns, fill matched rows.
    Returns (xlsx_bytes, matched, total)."""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(BytesIO(price_bytes), data_only=False)
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="Не удалось открыть файл «Прайс». Убедитесь, что это корректный XLS/XLSX.",
        )

    ws = wb.active
    header_row, art_col = _find_price_header_row(ws, _norm("Каталожный номер"))
    if header_row is None:
        wb.close()
        raise HTTPException(
            status_code=400,
            detail="В файле «Прайс» не найдена колонка «Каталожный номер».",
        )

    # Первые свободные колонки справа (1-based для openpyxl cell)
    tnved_col = ws.max_column + 1
    okpd_col = ws.max_column + 2

    ws.cell(row=header_row, column=tnved_col, value=tnved_header)
    ws.cell(row=header_row, column=okpd_col, value=okpd_header)

    matched = 0
    total = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        art_cell = ws.cell(row=row_idx, column=art_col + 1)
        key = _norm(art_cell.value)
        if not key:
            continue
        total += 1
        found = codes_map.get(key)
        if found is not None:
            matched += 1
            ws.cell(row=row_idx, column=tnved_col, value=found[0])
            ws.cell(row=row_idx, column=okpd_col, value=found[1])

    out = BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue(), matched, total


def codes_cache_path() -> str:
    """Локальный путь, куда сохраняется загруженная в систему «таблица с кодами»."""
    from app.config import settings as cfg
    return os.path.join(cfg.UPLOAD_DIR, "codes_saved.xlsx")


async def _resolve_codes(codes_file: Optional[UploadFile], db: AsyncSession) -> bytes:
    """Вернуть содержимое «таблицы с кодами».

    Если файл загружен вручную — используем его. Иначе берём ранее сохранённую
    в системе таблицу с кодами. Бросает HTTPException, если таблицы нет."""
    if codes_file is not None and codes_file.filename:
        data = await codes_file.read()
        if data:
            return data

    path = codes_cache_path()
    if not os.path.exists(path):
        raise HTTPException(
            status_code=400,
            detail="Таблица с кодами не загружена. Загрузите файл с ПК или сохраните его в системе.",
        )
    with open(path, "rb") as f:
        return f.read()


async def _resolve_price(price_file: Optional[UploadFile], db: AsyncSession) -> tuple[bytes, str]:
    """Вернуть (содержимое_прайса, базовое_имя).

    Если файл прайса загружен вручную — используем его. Иначе берём последний
    прайс, автоматически скачанный по FTP (раздел «Работа с таблицами» также
    поддерживает автозагрузку). Бросает HTTPException, если прайса нет."""
    if price_file is not None and price_file.filename:
        data = await price_file.read()
        if data:
            base = (price_file.filename or "прайс").rsplit(".", 1)[0]
            return data, base

    # Фолбэк: автоскачанный по FTP прайс.
    from app.services.ftp_service import price_cache_path
    from app.services.settings_service import get_setting

    path = price_cache_path()
    if not os.path.exists(path):
        raise HTTPException(
            status_code=400,
            detail="Прайс не загружен. Загрузите файл с ПК или настройте автозагрузку по FTP в настройках.",
        )
    with open(path, "rb") as f:
        data = f.read()
    last_name = (await get_setting(db, "price_ftp_last_name")) or "прайс_ftp.xlsx"
    base = last_name.rsplit(".", 1)[0]
    return data, base


async def _prepare_enriched(price_file: Optional[UploadFile], codes_file: Optional[UploadFile], db: AsyncSession):
    """Обработать прайс + таблицу кодов. Возвращает
    (xlsx_bytes, filename, matched, total, layout, codes_count)."""
    _patch_openpyxl_noneset()

    price_bytes, base_name = await _resolve_price(price_file, db)
    codes_bytes = await _resolve_codes(codes_file, db)

    loop = asyncio.get_event_loop()

    def _process():
        codes_map, tnved_header, okpd_header, layout = _build_codes_map(codes_bytes)
        xlsx, matched, total = _enrich_price(price_bytes, codes_map, tnved_header, okpd_header)
        return xlsx, matched, total, layout, len(codes_map)

    xlsx_data, matched, total, layout, codes_count = await loop.run_in_executor(None, _process)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{base_name}_с_кодами_{ts}.xlsx"
    return xlsx_data, filename, matched, total, layout, codes_count


@router.post("/enrich")
async def enrich_price_with_codes(
    request: Request,
    codes_file: Optional[UploadFile] = File(None),
    price_file: Optional[UploadFile] = File(None),
    db: AsyncSession = Depends(get_db),
):
    """Найти артикулы «Прайса» в «таблице с кодами», дописать колонки
    «код ТЭНВД» и «код ОКПД 2», вернуть готовый файл для скачивания.

    Прайс можно загрузить с ПК либо (если не приложен) взять автоскачанный по FTP.
    Таблицу с кодами можно загрузить с ПК либо взять ранее сохранённую в системе."""
    user = await get_current_user(request, db)
    if user.role.value not in ("admin", "staff"):
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    xlsx_data, filename, matched, total, layout, codes_count = await _prepare_enriched(price_file, codes_file, db)

    from urllib.parse import quote
    ascii_fallback = "price_with_codes.xlsx"
    cd = f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{quote(filename)}"

    return Response(
        content=xlsx_data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": cd,
            "X-Matched-Count": str(matched),
            "X-Total-Count": str(total),
            "X-Codes-Count": str(codes_count),
            "X-Codes-Layout": layout,
        },
    )


@router.post("/send")
async def enrich_and_send_email(
    request: Request,
    email: str = Form(...),
    subject: str = Form(""),
    codes_file: Optional[UploadFile] = File(None),
    price_file: Optional[UploadFile] = File(None),
    db: AsyncSession = Depends(get_db),
):
    """Обработать прайс + таблицу кодов и отправить результат письмом на указанный email."""
    user = await get_current_user(request, db)
    if user.role.value not in ("admin", "staff"):
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    to_email = (email or "").strip()
    if not to_email or "@" not in to_email:
        raise HTTPException(status_code=400, detail="Укажите корректный email получателя.")

    xlsx_data, filename, matched, total, layout, codes_count = await _prepare_enriched(price_file, codes_file, db)

    mail_subject = (subject or "").strip() or "Прайс с кодами ТН ВЭД и ОКПД 2"
    body = (
        f"Во вложении обработанный прайс.\n"
        f"Проставлено кодов: {matched} из {total} позиций. "
        f"Записей в таблице с кодами: {codes_count}."
    )

    from app.services.email_service import send_file_email
    try:
        await send_file_email(db, to_email, mail_subject, filename, xlsx_data, body_text=body)
    except RuntimeError as e:
        raise HTTPException(status_code=503, detail=str(e))
    except Exception as e:
        logger.error(f"Failed to send processed price to {to_email}: {e}", exc_info=True)
        raise HTTPException(status_code=502, detail=f"Не удалось отправить письмо: {e}")

    return JSONResponse({
        "ok": True,
        "message": f"Файл отправлен на {to_email}. Проставлено кодов: {matched} из {total}.",
        "matched": matched, "total": total,
    })


@router.get("/price/status")
async def price_ftp_status(request: Request, db: AsyncSession = Depends(get_db)):
    """Статус автоскачанного по FTP прайса: имя, дата, размер, настроен ли FTP."""
    user = await get_current_user(request, db)
    if user.role.value not in ("admin", "staff"):
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    from app.services.ftp_service import price_cache_path
    from app.services.settings_service import get_setting

    name = await get_setting(db, "price_ftp_last_name")
    at = await get_setting(db, "price_ftp_last_at")
    size = await get_setting(db, "price_ftp_last_size")
    folder = await get_setting(db, "ftp_price_path")
    host = await get_setting(db, "ftp_host")
    return JSONResponse({
        "available": bool(name) and os.path.exists(price_cache_path()),
        "filename": name or None,
        "downloaded_at": at or None,
        "size": int(size) if str(size).isdigit() else None,
        "folder": folder or None,
        "ftp_configured": bool(host),
    })


@router.post("/price/download")
async def price_ftp_download_now(request: Request, db: AsyncSession = Depends(get_db)):
    """Немедленно скачать прайс по FTP из настроенной папки и закешировать локально."""
    user = await get_current_user(request, db)
    if user.role.value not in ("admin", "staff"):
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    from app.services.ftp_service import download_and_cache_price
    try:
        res = await download_and_cache_price(db)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Manual FTP price download failed: {e}", exc_info=True)
        raise HTTPException(status_code=502, detail=f"Ошибка загрузки по FTP: {e}")

    return JSONResponse({
        "ok": True,
        "message": f"Прайс «{res['filename']}» загружен по FTP.",
        "filename": res["filename"],
        "size": res["size"],
    })


@router.get("/codes/status")
async def codes_saved_status(request: Request, db: AsyncSession = Depends(get_db)):
    """Статус сохранённой в системе «таблицы с кодами»: имя, дата, размер."""
    user = await get_current_user(request, db)
    if user.role.value not in ("admin", "staff"):
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    from app.services.settings_service import get_setting

    name = await get_setting(db, "codes_saved_name")
    at = await get_setting(db, "codes_saved_at")
    size = await get_setting(db, "codes_saved_size")
    return JSONResponse({
        "available": bool(name) and os.path.exists(codes_cache_path()),
        "filename": name or None,
        "saved_at": at or None,
        "size": int(size) if str(size).isdigit() else None,
    })


@router.post("/codes/save")
async def codes_save(
    request: Request,
    codes_file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Сохранить загруженную «таблицу с кодами» в системе для повторного использования."""
    user = await get_current_user(request, db)
    if user.role.value not in ("admin", "staff"):
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    if not codes_file or not codes_file.filename:
        raise HTTPException(status_code=400, detail="Выберите файл «таблица с кодами».")
    data = await codes_file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Файл пустой.")

    # Валидируем, что файл читается как таблица с кодами.
    _patch_openpyxl_noneset()
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, _build_codes_map, data)

    path = codes_cache_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "wb") as f:
        f.write(data)

    from datetime import timezone
    from app.services.settings_service import set_setting
    await set_setting(db, "codes_saved_name", codes_file.filename)
    await set_setting(db, "codes_saved_at", datetime.now(timezone.utc).isoformat())
    await set_setting(db, "codes_saved_size", str(len(data)))
    logger.info(f"Codes table saved in system: {codes_file.filename} ({len(data)} bytes)")

    return JSONResponse({
        "ok": True,
        "message": f"Таблица с кодами «{codes_file.filename}» сохранена.",
        "filename": codes_file.filename,
        "size": len(data),
    })


@router.delete("/codes")
async def codes_delete(request: Request, db: AsyncSession = Depends(get_db)):
    """Удалить сохранённую в системе «таблицу с кодами»."""
    user = await get_current_user(request, db)
    if user.role.value not in ("admin", "staff"):
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    path = codes_cache_path()
    if os.path.exists(path):
        try:
            os.remove(path)
        except OSError as e:
            logger.warning(f"Could not remove saved codes file: {e}")

    from app.services.settings_service import set_setting
    await set_setting(db, "codes_saved_name", "")
    await set_setting(db, "codes_saved_at", "")
    await set_setting(db, "codes_saved_size", "")

    return JSONResponse({"ok": True, "message": "Сохранённая таблица с кодами удалена."})
